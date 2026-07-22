#!/usr/bin/env python3
"""Export adhkar from Excel to SQL INSERT statements for webdev_execute_sql"""
import openpyxl
import json

wb = openpyxl.load_workbook('/home/ubuntu/upload/بياناتالجامعالمحررفيعملاليوموالليلة.xlsx', read_only=True)
ws = wb[wb.sheetnames[0]]

# Columns mapping:
# 0=الرقم, 1=كود الظرف, 2=الظرف, 3=نوع الظرف, 4=type_nl, 5=type_en
# 6=الوقت أو المكان أو الحال, 7=الذكر أو الدعاء بتمامه
# 8=translit, 9=text_nl, 10=text_en, 11=translation_status
# 12=العدد, 13=count_nl, 14=count_en
# 15=كيفية التطبيق, 16=kayfiyyah_nl, 17=kayfiyyah_en
# 18=الفضل والأجر, 19=fadl_nl, 20=fadl_en
# 21=المرتبة, 22=martaba_nl, 23=martaba_en
# 24=القسم, 25=section_nl, 26=section_en
# 27=الباب, 28=chapter_nl, 29=chapter_en

time_contexts = {'morning', 'evening', 'sleep', 'waking', 'pre_fajr', 'fajr_period', 
                 'duha_work', 'dhuhr_time', 'asr_time', 'maghrib_isha', 'night_prayer',
                 'sleep_cycle', 'witr'}
event_contexts = {'adhan', 'after_every_prayer', 'after_fajr', 'after_maghrib', 
                  'inside_prayer', 'special_prayers', 'wudu', 'mosque', 'friday',
                  'hajj', 'fasting', 'zakat', 'monthly', 'yearly', 'arafah',
                  'new_moon', 'eclipse', 'khutbah', 'quran_duas', 'recitation',
                  'sahabah_duas', 'istikharah'}
state_contexts = {'anger', 'distress', 'difficulty', 'joy', 'gratitude', 'poverty',
                  'debt', 'loan_repay', 'enemy_fear', 'waswas', 'evil_eye',
                  'pain_ruqyah', 'sick_visit', 'death_funeral', 'hidden_shirk',
                  'sin_majlis', 'seeing_afflicted', 'love_in_allah', 'omens',
                  'night_fright', 'stings', 'epidemic', 'drought', 'floods', 'weather',
                  'animal_sounds', 'menses'}
place_contexts = {'home', 'market', 'toilet', 'graveyard', 'entering_town', 'travel',
                  'travel_route', 'mosque'}
life_contexts = {'marriage', 'newborn', 'slaughter', 'food', 'clothing', 'sneezing',
                 'first_fruits'}

def get_category(ctx):
    if ctx in time_contexts: return 'time'
    if ctx in event_contexts: return 'event'
    if ctx in state_contexts: return 'state'
    if ctx in place_contexts: return 'place'
    if ctx in life_contexts: return 'life'
    return 'general'

def escape_sql(s):
    if not s:
        return ''
    return str(s).replace("'", "''").replace("\\", "\\\\")

# Write SQL in batches of 20
batch_size = 20
batch_num = 0
rows_data = []

for row in ws.iter_rows(min_row=2, values_only=True):
    if not row[0]:
        continue
    rows_data.append(row)

wb.close()

total_batches = (len(rows_data) + batch_size - 1) // batch_size
print(f"Total: {len(rows_data)} adhkar, {total_batches} batches")

for batch_idx in range(total_batches):
    start = batch_idx * batch_size
    end = min(start + batch_size, len(rows_data))
    batch = rows_data[start:end]
    
    sql_parts = []
    for i, row in enumerate(batch):
        context = escape_sql(row[1]) if row[1] else 'general'
        category = get_category(str(row[1]) if row[1] else 'general')
        text_ar = escape_sql(row[7])  # الذكر أو الدعاء بتمامه
        text_nl = escape_sql(row[9])  # text_nl
        text_en = escape_sql(row[10])  # text_en
        how_ar = escape_sql(row[15])  # كيفية التطبيق
        how_nl = escape_sql(row[16])  # kayfiyyah_nl
        how_en = escape_sql(row[17])  # kayfiyyah_en
        reward_ar = escape_sql(row[18])  # الفضل والأجر
        reward_nl = escape_sql(row[19])  # fadl_nl
        reward_en = escape_sql(row[20])  # fadl_en
        guidance_ar = escape_sql(row[6])  # الوقت أو المكان أو الحال (context description)
        guidance_nl = escape_sql(row[4]) if row[4] else ''  # type_nl as context description
        guidance_en = escape_sql(row[5]) if row[5] else ''  # type_en as context description
        reps = int(row[12]) if row[12] and str(row[12]).isdigit() else 1
        sort_order = start + i + 1
        
        sql_parts.append(f"('{context}', '{category}', '{text_ar}', '{text_nl}', '{text_en}', '{how_ar}', '{how_nl}', '{how_en}', '{reward_ar}', '{reward_nl}', '{reward_en}', '{guidance_ar}', '{guidance_nl}', '{guidance_en}', {reps}, {sort_order})")
    
    sql = f"""INSERT INTO adhkar (context_code, category, text_ar, text_nl, text_en, how_to_apply_ar, how_to_apply_nl, how_to_apply_en, reward_ar, reward_nl, reward_en, guidance_ar, guidance_nl, guidance_en, repetitions, sort_order) VALUES\n{','.join(sql_parts)};"""
    
    # Write each batch to a file
    with open(f'/home/ubuntu/opvoedadvies_apk/scripts/adhkar_batch_{batch_idx:02d}.sql', 'w') as f:
        f.write(sql)

print(f"Generated {total_batches} SQL batch files")
