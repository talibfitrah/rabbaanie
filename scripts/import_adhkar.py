#!/usr/bin/env python3
"""Import adhkar from Excel to database"""
import openpyxl
import psycopg2
import os

DB_URL = os.environ.get("DATABASE_URL", "")
# Parse from .env file
if not DB_URL:
    env_path = "/home/ubuntu/opvoedadvies_apk/.env"
    if os.path.exists(env_path):
        with open(env_path) as f:
            for line in f:
                if line.startswith("DATABASE_URL="):
                    DB_URL = line.strip().split("=", 1)[1].strip('"').strip("'")
                    break

if not DB_URL:
    # Try reading from project config
    import json
    config_path = "/home/ubuntu/opvoedadvies_apk/.project-config.json"
    if os.path.exists(config_path):
        with open(config_path) as f:
            config = json.load(f)
            DB_URL = config.get("database", {}).get("url", "")

print(f"DB URL found: {'Yes' if DB_URL else 'No'}")

# Read Excel
wb = openpyxl.load_workbook('/home/ubuntu/upload/بياناتالجامعالمحررفيعملاليوموالليلة.xlsx', read_only=True)
ws = wb[wb.sheetnames[0]]

# Get headers
headers = []
for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
    headers = [str(h) if h else f"col_{i}" for i, h in enumerate(row)]
    break

print(f"Headers: {headers[:10]}")

# Collect all rows
rows_data = []
for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
    if not row[0] and not row[1]:
        continue
    rows_data.append(row)

print(f"Total adhkar to import: {len(rows_data)}")

# Map context categories
time_contexts = {'morning', 'evening', 'sleep', 'waking', 'pre_fajr', 'fajr_period', 
                 'duha_work', 'dhuhr_time', 'asr_time', 'maghrib_isha', 'night_prayer',
                 'sleep_cycle', 'witr'}
event_contexts = {'adhan', 'after_every_prayer', 'after_fajr', 'after_maghrib', 
                  'inside_prayer', 'special_prayers', 'wudu', 'mosque', 'friday',
                  'hajj', 'fasting', 'zakat', 'monthly', 'yearly', 'arafah',
                  'new_moon', 'eclipse', 'khutbah', 'quran_duas', 'recitation'}
state_contexts = {'anger', 'distress', 'difficulty', 'joy', 'gratitude', 'poverty',
                  'debt', 'loan_repay', 'enemy_fear', 'waswas', 'evil_eye',
                  'pain_ruqyah', 'sick_visit', 'death_funeral', 'hidden_shirk',
                  'sin_majlis', 'seeing_afflicted', 'love_in_allah', 'omens',
                  'night_fright', 'stings', 'epidemic', 'drought', 'floods', 'weather'}

def get_category(ctx):
    if ctx in time_contexts: return 'time'
    if ctx in event_contexts: return 'event'
    if ctx in state_contexts: return 'state'
    return 'general'

# Connect and insert
conn = psycopg2.connect(DB_URL + "?sslmode=require")
cur = conn.cursor()

# Clear existing
cur.execute("DELETE FROM adhkar")

inserted = 0
for i, row in enumerate(rows_data):
    # Columns: 0=id, 1=context, 2=text_ar, 3=text_nl, 4=text_en, 
    # 5=howToApply_ar, 6=howToApply_nl, 7=howToApply_en,
    # 8=reward_ar, 9=reward_nl, 10=reward_en,
    # 11=guidance_ar, 12=guidance_nl, 13=guidance_en, 14=repetitions
    context = str(row[1]) if row[1] else 'general'
    text_ar = str(row[2]) if row[2] else ''
    text_nl = str(row[3]) if row[3] else ''
    text_en = str(row[4]) if row[4] else ''
    how_ar = str(row[5]) if len(row) > 5 and row[5] else ''
    how_nl = str(row[6]) if len(row) > 6 and row[6] else ''
    how_en = str(row[7]) if len(row) > 7 and row[7] else ''
    reward_ar = str(row[8]) if len(row) > 8 and row[8] else ''
    reward_nl = str(row[9]) if len(row) > 9 and row[9] else ''
    reward_en = str(row[10]) if len(row) > 10 and row[10] else ''
    guidance_ar = str(row[11]) if len(row) > 11 and row[11] else ''
    guidance_nl = str(row[12]) if len(row) > 12 and row[12] else ''
    guidance_en = str(row[13]) if len(row) > 13 and row[13] else ''
    reps = int(row[14]) if len(row) > 14 and row[14] else 1
    
    if not text_ar:
        continue
    
    category = get_category(context)
    
    cur.execute("""
        INSERT INTO adhkar (context_code, category, text_ar, text_nl, text_en,
            how_to_apply_ar, how_to_apply_nl, how_to_apply_en,
            reward_ar, reward_nl, reward_en,
            guidance_ar, guidance_nl, guidance_en,
            repetitions, sort_order)
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
    """, (context, category, text_ar, text_nl, text_en,
          how_ar, how_nl, how_en, reward_ar, reward_nl, reward_en,
          guidance_ar, guidance_nl, guidance_en, reps, i))
    inserted += 1

conn.commit()
cur.close()
conn.close()

print(f"Successfully imported {inserted} adhkar!")
